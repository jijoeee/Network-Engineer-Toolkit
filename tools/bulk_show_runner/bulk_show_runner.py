#!/usr/bin/env python3
"""
Bulk Show-Command Runner — Network Engineer Toolkit
===================================================

Runs the SAME list of `show` commands across every device in your inventory,
in parallel, and drops every answer into ONE Excel workbook — one sheet per
command, one row per device. Run it again after a change and it tells you which
rows are different AND what the difference is, so a pre/post-change comparison
is one click instead of an evening of pasting screenshots into Word.

Reuses the same `devices.xlsx` layout as the Bulk Ping Monitor and the Config
Backup Runner (Hostname, IP Address, Location, Device Type, Group), so you keep
one device list for the whole toolkit.

What it does, in order:
  1. Reads the device list from `devices.xlsx`.
  2. Reads the commands to run from `commands.txt` (one command per line). If
     that file is missing it asks you to type them in.
  3. Pings each device first, so dead boxes are skipped instead of blocking on a
     30-second SSH timeout.
  4. Logs in over SSH (Netmiko) and runs every command on every device.
     Vendor is taken from the "Device Type" column, exactly like the Config
     Backup Runner (cisco / cisco_xe / huawei / hp_comware / arista / juniper).
  5. Writes ONE workbook to `report/Show_Report_YYYYMMDD_HHMM.xlsx`:
       - a "Summary" sheet: one row per device, OK/FAILED and why
       - one sheet per command: one row per device with that command's output
  6. Saves a snapshot to `snapshots/` so a later run can be compared.
  7. If a previous snapshot exists, each command sheet gains two columns —
     "Changed?" (CHANGED / SAME / NEW) and "What changed" (the actual added and
     removed lines) — and changed rows are coloured amber. That is your
     before/after health check.

Credentials are asked for at run time, or read from the environment
(NET_USER / NET_PASS / NET_ENABLE). They are never written to a file, and they
are masked out of any error message before it reaches the screen or the report.

Nothing here pushes config to a device — it only runs the commands you put in
`commands.txt`, so it is read-only by design.

--------------------------------------------------------------------------------
HOW TO USE
--------------------------------------------------------------------------------
    pip install -r requirements.txt          # netmiko, openpyxl
    cd tools/bulk_show_runner

    # 1. Put your device list in devices.xlsx (a sample ships in this folder):
    #    Hostname | IP Address | Location | Device Type (vendor) | Group
    #    Device Type must name the VENDOR, not the form factor.

    # 2. Put the commands you want to run in commands.txt, one per line, e.g.:
    #        show ip interface brief
    #        show cdp neighbors
    #        show version

    # 3. Run it:
    python bulk_show_runner.py

    # 4. Open report/Show_Report_*.xlsx — one tab per command.

    # 5. Make your change, run it again. The new workbook marks every row whose
    #    output is different from last time and shows you the changed lines.

Optional flags:
    python bulk_show_runner.py -c mycommands.txt   # use a different command file
    python bulk_show_runner.py --no-compare        # skip the before/after diff
--------------------------------------------------------------------------------
"""

import os
import re
import sys
import glob
import json
import time
import getpass
import difflib
import argparse
import platform
import subprocess
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# --- Dependencies -----------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl is required -> pip install openpyxl")
    sys.exit(1)

try:
    from netmiko import ConnectHandler
    from netmiko.exceptions import (
        NetmikoAuthenticationException,
        NetmikoTimeoutException,
    )
except ImportError:
    print("ERROR: netmiko is required -> pip install netmiko")
    sys.exit(1)


# --- Path resolution (matches the rest of the toolkit) ----------------------
def get_base_dir():
    """Folder the script/exe lives in, so it works both as .py and frozen .exe."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


BASE_DIR = get_base_dir()
EXCEL_FILE = os.path.join(BASE_DIR, "devices.xlsx")
COMMANDS_FILE = os.path.join(BASE_DIR, "commands.txt")
SNAPSHOT_DIR = os.path.join(BASE_DIR, "snapshots")

# SSH sessions are heavier than pings, and a TACACS/RADIUS back-end dislikes a
# stampede — keep the pool modest.
MAX_WORKERS = 10
PING_TIMEOUT_S = 2

# Map the human "Device Type" cell to a Netmiko device_type. Same vendors and
# the same matching rules as the Config Backup Runner, so one inventory drives
# the whole toolkit.
VENDOR_MAP = {
    "cisco":        "cisco_ios",
    "cisco_ios":    "cisco_ios",
    "ios":          "cisco_ios",
    "cisco_xe":     "cisco_xe",
    "cisco_ios_xe": "cisco_xe",
    "ios_xe":       "cisco_xe",
    "iosxe":        "cisco_xe",
    "huawei":       "huawei",
    "hp_comware":   "hp_comware",
    "arista":       "arista_eos",
    "arista_eos":   "arista_eos",
    "juniper":      "juniper_junos",
}
DEFAULT_VENDOR = "cisco_ios"

# Longest keys first, so "arista_eos" is tested before "arista". Matching is
# done inside the cell, not on the whole cell, because real inventories say
# "Huawei S5720" or "Cisco IOS-XE router", not "huawei".
#
# Longest-first is why "cisco_ios_xe" and "ios_xe" have to be keys of their own:
# normalised, "Cisco IOS-XE router" is "cisco_ios_xe_router", which contains
# BOTH "cisco_ios" (9) and "cisco_xe" (8), and the longer of those two would
# otherwise win and hand an IOS-XE box the plain cisco_ios driver. The
# 12-character key wins instead. Same bug, same fix, as the Config Backup
# Runner — do not "simplify" this back to an exact-match lookup.
VENDOR_KEYS_BY_LENGTH = sorted(VENDOR_MAP, key=len, reverse=True)


def resolve_vendor(device_type):
    """Return (netmiko_device_type, was_defaulted).

    The cell is matched vendor-first: an exact hit wins, otherwise the first
    recognised vendor name appearing anywhere in the cell wins. Form-factor
    words on their own ("router", "switch", "firewall") deliberately do NOT
    match — they name the shape of the box, not the CLI it speaks — so they
    fall back to cisco_ios and are reported as a fallback rather than silently
    treated as Cisco.
    """
    key = (device_type or "").strip().lower().replace(" ", "_").replace("-", "_")
    if key in VENDOR_MAP:
        return VENDOR_MAP[key], False

    for vendor in VENDOR_KEYS_BY_LENGTH:
        if vendor in key:
            return VENDOR_MAP[vendor], False

    return DEFAULT_VENDOR, True


# --- Inventory --------------------------------------------------------------
def load_devices():
    """Read devices.xlsx exactly like the Ping Monitor does (row 1 = header)."""
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: could not find '{EXCEL_FILE}'")
        print("Put a devices.xlsx next to this script with columns: "
              "Hostname, IP Address, Location, Device Type, Group")
        sys.exit(1)

    devices = []
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"ERROR: '{EXCEL_FILE}' could not be read as an Excel workbook.")
        print(f"       {type(e).__name__}: {e}")
        print("       It must be a real .xlsx saved by Excel or LibreOffice — "
              "a renamed .csv or .xls will not work.")
        sys.exit(1)

    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        devices.append({
            "hostname":    str(row[0]).strip(),
            "ip":          str(row[1]).strip(),
            "location":    str(row[2]).strip() if len(row) > 2 and row[2] else "-",
            "device_type": str(row[3]).strip() if len(row) > 3 and row[3] else "-",
            "group":       str(row[4]).strip() if len(row) > 4 and row[4] else "-",
        })
    return devices


# --- Commands ---------------------------------------------------------------
def load_commands(commands_path):
    """One command per line. Lines starting with # are comments; blanks ignored."""
    commands = []
    if os.path.exists(commands_path):
        with open(commands_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    commands.append(line)
        if commands:
            return commands
        print(f"'{os.path.basename(commands_path)}' has no commands in it "
              "(every line is blank or commented out).")

    # No file (or an empty one) — ask for the commands interactively.
    print("Type the show commands to run, one per line. Blank line to finish.")
    print("(Tip: put them in commands.txt next time so you don't retype them.)")
    while True:
        try:
            line = input("  cmd> ").strip()
        except EOFError:
            # Scheduled/unattended run with no terminal to type into.
            print()
            print("ERROR: no commands to run and nothing to type into "
                  "(no interactive terminal).")
            print(f"       Put one command per line in "
                  f"'{os.path.basename(commands_path)}' and run it again.")
            sys.exit(1)
        if not line:
            break
        commands.append(line)
    return commands


# --- Credential hygiene -----------------------------------------------------
def scrub(text, creds):
    """Mask the password/enable secret if they ever appear in text we log."""
    for key in ("password", "secret"):
        value = creds.get(key)
        if value and value in text:
            text = text.replace(value, "***")
    return text


# --- Reachability -----------------------------------------------------------
def is_reachable(ip):
    """One cheap ping so we skip dead devices instead of waiting on SSH timeouts."""
    if platform.system().lower() == "windows":
        cmd = ["ping", "-n", "1", "-w", str(PING_TIMEOUT_S * 1000), ip]
    else:
        cmd = ["ping", "-c", "1", "-W", str(PING_TIMEOUT_S), ip]
    try:
        completed = subprocess.run(
            cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
        )
        return completed.returncode == 0
    except Exception:
        return False


# --- Run the commands on one device -----------------------------------------
def run_device(device, creds, commands):
    """Return {hostname, ip, device_type, status, detail, outputs:{cmd->text}}."""
    hostname = device["hostname"]
    ip = device["ip"]
    result = {
        "hostname": hostname, "ip": ip, "location": device["location"],
        "device_type": device["device_type"], "group": device["group"],
        "status": "FAILED", "detail": "", "outputs": {},
    }

    if not is_reachable(ip):
        result["detail"] = "unreachable (ping failed) — skipped"
        return result

    dt, defaulted = resolve_vendor(device["device_type"])
    if defaulted:
        result["detail"] = f"unknown device type '{device['device_type']}', assumed cisco_ios; "

    conn_params = {
        "device_type": dt,
        "host": ip,
        "username": creds["username"],
        "password": creds["password"],
        "conn_timeout": 15,
        "fast_cli": False,
    }
    if creds.get("secret"):
        conn_params["secret"] = creds["secret"]

    try:
        errored = 0
        with ConnectHandler(**conn_params) as conn:
            if creds.get("secret"):
                try:
                    conn.enable()
                except Exception:
                    pass  # not all platforms/accounts need enable
            for cmd in commands:
                try:
                    out = conn.send_command(cmd, read_timeout=60)
                    result["outputs"][cmd] = out if out is not None else ""
                except Exception as e:
                    # Written into the workbook, so it gets the same masking as
                    # the connection errors below.
                    result["outputs"][cmd] = f"<command error: {scrub(str(e), creds)}>"
                    errored += 1

        ran = len(result["outputs"])
        result["status"] = "OK"
        result["detail"] += f"{ran} command(s) run"
        if errored:
            result["detail"] += f", {errored} errored"
        return result

    except NetmikoAuthenticationException:
        result["detail"] += "authentication failed"
    except NetmikoTimeoutException:
        result["detail"] += "connection timed out"
    except Exception as e:
        # This string is printed AND written to the workbook. A third-party
        # library's exception text is not ours to trust, so anything that looks
        # like the credentials we just handed it is masked before it persists.
        result["detail"] += f"error: {scrub(str(e), creds)}"
    return result


# --- Snapshots (for the before/after diff) ----------------------------------
def save_snapshot(results, commands):
    """Snapshot the devices that actually answered.

    FAILED devices are deliberately left out. If a whole run fails — one wrong
    password, one unreachable management VLAN — a snapshot full of empty
    outputs would become the baseline, and the next good run would report every
    device as NEW. The last useful baseline is worth more than the newest one.
    """
    keepers = [r for r in results if r["status"] == "OK"]
    if not keepers:
        return None

    os.makedirs(SNAPSHOT_DIR, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    path = os.path.join(SNAPSHOT_DIR, f"snapshot_{stamp}.json")
    payload = {
        "taken": datetime.now().isoformat(timespec="seconds"),
        "commands": commands,
        "devices": {
            r["hostname"]: {
                "status": r["status"],
                "outputs": r["outputs"],
            } for r in keepers
        },
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    return path


def load_previous_snapshot():
    """Most recent snapshot *before* the one we're about to write, or None."""
    if not os.path.isdir(SNAPSHOT_DIR):
        return None, None
    files = sorted(glob.glob(os.path.join(SNAPSHOT_DIR, "snapshot_*.json")))
    if not files:
        return None, None
    latest = files[-1]
    try:
        with open(latest, "r", encoding="utf-8") as f:
            return json.load(f), os.path.basename(latest)
    except Exception:
        return None, None


def _normalise(text):
    """Ignore trailing whitespace so cosmetic differences don't flag as CHANGED."""
    if text is None:
        return ""
    lines = [ln.rstrip() for ln in str(text).splitlines()]
    while lines and not lines[-1]:
        lines.pop()
    return "\n".join(lines)


def compare_output(prev_snapshot, hostname, cmd, current):
    """Return (verdict, previous_text) for one device+command.

    verdict is CHANGED / SAME / NEW, or "" when there is nothing to compare to.
    """
    if not prev_snapshot:
        return "", None
    dev = prev_snapshot.get("devices", {}).get(hostname)
    if not dev or cmd not in dev.get("outputs", {}):
        return "NEW", None
    previous = dev["outputs"][cmd]
    if _normalise(previous) == _normalise(current):
        return "SAME", previous
    return "CHANGED", previous


def describe_change(previous, current, max_lines=12):
    """The point of the whole tool: not *that* it changed, but *what* changed.

    Returns a short human summary — a count, then the removed and added lines,
    capped so one chatty device cannot produce an unreadable cell.
    """
    prev_lines = _normalise(previous).splitlines()
    curr_lines = _normalise(current).splitlines()

    added, removed = [], []
    for line in difflib.unified_diff(prev_lines, curr_lines, lineterm="", n=0):
        if line.startswith(("+++", "---", "@@")):
            continue
        if line.startswith("+"):
            added.append(line)
        elif line.startswith("-"):
            removed.append(line)

    if not added and not removed:
        return ""

    shown = (removed + added)[:max_lines]
    hidden = len(removed) + len(added) - len(shown)
    body = "\n".join(shown)
    if hidden > 0:
        body += f"\n...and {hidden} more changed line(s)"
    return f"-{len(removed)} / +{len(added)} line(s)\n{body}"


# --- Workbook ---------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
CHANGED_FILL = PatternFill("solid", fgColor="FFE699")   # amber
NEW_FILL = PatternFill("solid", fgColor="D9E1F2")       # pale blue
TOP_ALIGN = Alignment(vertical="top", wrap_text=True)

# Excel refuses more than 32,767 characters in a cell, and openpyxl refuses the
# control characters a pager or a banner can leave in CLI output. Either one
# raises *while saving*, i.e. after every device has already been logged into —
# so both are cleaned on the way in, not discovered on the way out.
EXCEL_CELL_LIMIT = 32767
ANSI_RE = re.compile(r"\x1b\[[0-9;?]*[ -/]*[@-~]")
CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
TRUNCATED_NOTE = ("\n...[truncated: output is longer than Excel's "
                  "32,767-character cell limit — the full text is in snapshots/]")


def _cell_safe(text):
    """Make device output safe to put in a cell without losing the whole run."""
    if text is None:
        return ""
    cleaned = CONTROL_RE.sub("", ANSI_RE.sub("", str(text)))
    if len(cleaned) > EXCEL_CELL_LIMIT:
        cleaned = cleaned[:EXCEL_CELL_LIMIT - len(TRUNCATED_NOTE)] + TRUNCATED_NOTE
    return cleaned


def _safe_sheet_title(name, used):
    """Excel sheet names: <=31 chars, no []:*?/\\ , and must be unique."""
    title = re.sub(r"[\[\]:*?/\\]", " ", name).strip().strip("'")[:31] or "cmd"
    base, n = title, 1
    while title.lower() in used:
        suffix = f" ({n})"
        title = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center")


def write_workbook(results, commands, prev_snapshot, prev_name):
    report_dir = os.path.join(BASE_DIR, "report")
    os.makedirs(report_dir, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M")
    path = os.path.join(report_dir, f"Show_Report_{stamp}.xlsx")

    wb = openpyxl.Workbook()
    used_titles = set()
    changed_per_device = {r["hostname"]: 0 for r in results}

    # --- One sheet per command ---------------------------------------------
    # Built before the Summary is filled in, because the Summary reports how
    # many commands changed per device and that is only known once they are.
    command_sheets = []
    for cmd in commands:
        ws = wb.create_sheet(_safe_sheet_title(cmd, used_titles))
        headers = ["Hostname", "IP Address", "Device Type", "Status"]
        if prev_snapshot:
            headers += ["Changed?", "What changed"]
        headers.append("Output")
        ws.append(headers)
        _style_header(ws, len(headers))

        out_col = len(headers)

        for r in results:
            output = r["outputs"].get(
                cmd, "" if r["status"] == "OK" else "(not run — device failed)")
            row = [r["hostname"], r["ip"], r["device_type"], r["status"]]

            change = ""
            if prev_snapshot:
                previous = None
                if r["status"] == "OK":
                    change, previous = compare_output(
                        prev_snapshot, r["hostname"], cmd, output)
                detail = ""
                if change == "CHANGED":
                    changed_per_device[r["hostname"]] += 1
                    detail = describe_change(previous, output)
                row += [change, _cell_safe(detail)]

            row.append(_cell_safe(output))
            ws.append(row)

            excel_row = ws.max_row
            ws.cell(row=excel_row, column=out_col).alignment = TOP_ALIGN
            if prev_snapshot:
                ws.cell(row=excel_row, column=6).alignment = TOP_ALIGN
                if change == "CHANGED":
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=excel_row, column=col).fill = CHANGED_FILL
                elif change == "NEW":
                    ws.cell(row=excel_row, column=5).fill = NEW_FILL

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 10
        if prev_snapshot:
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 50
        ws.column_dimensions[openpyxl.utils.get_column_letter(out_col)].width = 90
        ws.freeze_panes = "A2"
        command_sheets.append(ws)

    # --- Summary sheet, moved to the front ----------------------------------
    summary = wb.active                      # the empty default sheet
    summary.title = "Summary"
    headers = ["Hostname", "IP Address", "Location", "Device Type",
               "Group", "Status", "Detail"]
    if prev_snapshot:
        headers.append("Commands changed")
    summary.append(headers)
    for r in results:
        row = [r["hostname"], r["ip"], r["location"], r["device_type"],
               r["group"], r["status"], r["detail"]]
        if prev_snapshot:
            row.append(changed_per_device.get(r["hostname"], 0))
        summary.append(row)
    _style_header(summary, len(headers))

    widths = [20, 16, 14, 14, 12, 10, 40, 18]
    for i, w in enumerate(widths[:len(headers)], start=1):
        summary.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    summary.freeze_panes = "A2"
    if prev_name:
        note = summary.cell(
            row=len(results) + 3, column=1,
            value=f"Changed? column compares each command against: {prev_name}")
        note.font = Font(italic=True, color="808080")
    wb.move_sheet(summary, offset=-len(command_sheets))

    wb.save(path)
    return path


# --- Credentials ------------------------------------------------------------
def get_credentials():
    """From environment if set, otherwise prompt. Never written to disk."""
    username = os.environ.get("NET_USER")
    password = os.environ.get("NET_PASS")
    secret = os.environ.get("NET_ENABLE", "")

    if not username:
        username = input("SSH username: ").strip()
    if not password:
        password = getpass.getpass("SSH password: ")
    if not secret:
        secret = getpass.getpass("Enable secret (blank if none): ")

    return {"username": username, "password": password, "secret": secret}


# --- Main -------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Run the same show commands across every device in devices.xlsx.")
    parser.add_argument("-c", "--commands", default=COMMANDS_FILE,
                        help="path to the command list (default: commands.txt)")
    parser.add_argument("--no-compare", action="store_true",
                        help="do not compare against the previous run")
    args = parser.parse_args()

    print("=" * 60)
    print(" Bulk Show-Command Runner — Network Engineer Toolkit")
    print("=" * 60)

    devices = load_devices()
    if not devices:
        print("No devices found in devices.xlsx. Nothing to do.")
        return
    print(f"Loaded {len(devices)} device(s) from {os.path.basename(EXCEL_FILE)}.")

    commands = load_commands(args.commands)
    if not commands:
        print("No commands to run. Nothing to do.")
        return
    print(f"Running {len(commands)} command(s) per device:")
    for cmd in commands:
        print(f"    - {cmd}")

    # Say which devices we could not identify a vendor for BEFORE the run, not
    # after. On a 200-device inventory that is the difference between fixing one
    # spreadsheet column and reading 200 FAILED rows.
    unknown = [d for d in devices if resolve_vendor(d["device_type"])[1]]
    if unknown:
        print()
        print(f"WARNING: {len(unknown)} of {len(devices)} device(s) have no recognised "
              f"vendor in the 'Device Type' column.")
        print(f"         They will be tried as {DEFAULT_VENDOR}, which will fail "
              f"on a non-Cisco box.")
        for d in unknown[:10]:
            print(f"         - {d['hostname']:<20} Device Type = '{d['device_type']}'")
        if len(unknown) > 10:
            print(f"         - ...and {len(unknown) - 10} more")
        print("         Put the vendor in that column (cisco / cisco_xe / huawei / "
              "arista / juniper / hp_comware).")
        print()

    # Load the previous snapshot BEFORE we write this run's snapshot.
    prev_snapshot, prev_name = (None, None)
    if not args.no_compare:
        prev_snapshot, prev_name = load_previous_snapshot()
        if prev_name:
            print(f"Comparing against previous run: {prev_name}")

    creds = get_credentials()
    print("-" * 60)

    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {
            pool.submit(run_device, d, creds, commands): d for d in devices
        }
        for fut in as_completed(futures):
            r = fut.result()
            mark = "OK  " if r["status"] == "OK" else "FAIL"
            print(f"[{mark}] {r['hostname']:<20} {r['ip']:<16} {r['detail']}")
            results.append(r)

    # Keep the on-screen order stable (as_completed returns out of order).
    order = {d["hostname"]: i for i, d in enumerate(devices)}
    results.sort(key=lambda r: order.get(r["hostname"], 0))

    ok = sum(1 for r in results if r["status"] == "OK")
    failed = len(results) - ok
    print("-" * 60)
    print(f"Done: {ok} OK, {failed} failed.")

    report_path = write_workbook(results, commands, prev_snapshot, prev_name)
    print(f"Workbook -> {report_path}")

    snap_path = save_snapshot(results, commands)
    if snap_path:
        print(f"Snapshot -> {snap_path}")
    else:
        print("No device answered, so no snapshot was saved — "
              "your last good baseline is kept for the next run.")
    if prev_name:
        print("Rows highlighted amber in the workbook changed since the last run; "
              "the 'What changed' column says how.")


if __name__ == "__main__":
    main()
